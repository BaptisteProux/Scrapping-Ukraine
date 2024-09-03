import requests as r
from bs4 import BeautifulSoup as bs
import pandas as pd
import re
from itertools import chain
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os
from os.path import exists
import sys
from datetime import datetime, date, timedelta
import xlsxwriter as xs
from openpyxl.workbook.child import INVALID_TITLE_REGEX
from openpyxl import load_workbook
import time
from requests.exceptions import SSLError
from requests.exceptions import ConnectionError
import glob


###########################################################################################
def parser(url,string="",classe=""):
	page=r.get(url)
	parser=bs(page.content,"html.parser")
	parser.encoding = parser.apparent_encoding
	if string=="" :
		return parser
	else:
		if classe=="":
			sous_section=parser.find_all(str(string))
		else:
			sous_section=parser.find_all(str(string),{"class": classe})
		return sous_section

def texte(data,split):
	text_all=[]
	for i in range(len(data)):
		text=data[i].get_text().split(split)
		text_all.append(text)
	text_all=list(chain.from_iterable(text_all))
	return text_all


def convert_date_to_string(string_date):
	date = datetime.strftime(string_date, '%d.%m.%Y')
	return date

def convert_string_to_date(date):
	string_date = datetime.strptime(date, '%d.%m.%Y').date()
	return string_date

def excel_liveuamap(string_start_date,diff,url):
	i=0;

	url_date=url+string_start_date
	#print(url_date)
	sous_liveuamap_text=parser(url_date,"div","title")
	sous_liveuamap_when=parser(url_date,"span","date_add")

	liveuamap_text=texte(sous_liveuamap_text,'<div class="title">')
	liveuamap_date=texte(sous_liveuamap_when,'<span class="date_add">')
	liveuamap_date=liveuamap_date[1:]
	df_liveuamap=pd.DataFrame(liveuamap_text,index=liveuamap_date)

	df_liveuamap=pd.concat([df_liveuamap],keys=['liveuamap'+string_start_date])

	if exists(file_excel)!=True:
		with pd.ExcelWriter(file_excel) as writer:
			df_liveuamap.to_excel(writer, sheet_name=string_start_date)	

	xls = pd.ExcelFile(file_excel)
	sheets = xls.sheet_names

	while i<=diff:				

		if string_start_date not in sheets or i==diff: 
			url_date=url+string_start_date
			print(url_date)
			sous_liveuamap_text=parser(url_date,"div","title")
			time.sleep(1)
			sous_liveuamap_when=parser(url_date,"span","date_add")

			liveuamap_text=texte(sous_liveuamap_text,'<div class="title">')
			liveuamap_date=texte(sous_liveuamap_when,'<span class="date_add">')
			liveuamap_date=liveuamap_date[1:]
			df_liveuamap=pd.DataFrame(liveuamap_text,index=liveuamap_date)
			#print(df_liveuamap)

			df_liveuamap=pd.concat([df_liveuamap],keys=['liveuamap'+string_start_date])

			if string_start_date not in sheets:
				workbook = openpyxl.load_workbook(file_excel)
				workbook.create_sheet(index=0, title=string_start_date)
				workbook.save(file_excel)

			with pd.ExcelWriter(file_excel,engine="openpyxl",mode="a", if_sheet_exists="replace") as writer:
				df_liveuamap.to_excel(writer, sheet_name=string_start_date)
		else:
			print("Déjà dans le excel")



		start_date=convert_string_to_date(string_start_date)
		start_date= start_date + timedelta(days=1)

		string_start_date=convert_date_to_string(start_date)

		i=i+1

	os.system('"'+file_excel+ '"')



def file_liveuamap(string_start_date,diff,url):
	i=0;
	while i<=diff:
		url_date=url+string_start_date
		print(url_date)
		sous_liveuamap_text=parser(url_date,"div","title")
		time.sleep(1)
		sous_liveuamap_when=parser(url_date,"span","date_add")

		liveuamap_text=texte(sous_liveuamap_text,'<div class="title">')
		liveuamap_date=texte(sous_liveuamap_when,'<span class="date_add">')
		liveuamap_date=liveuamap_date[1:]
		df_liveuamap=pd.DataFrame(liveuamap_text,index=liveuamap_date)

		df_liveuamap=pd.concat([df_liveuamap],keys=['liveuamap'+string_start_date])

		with open(file_txt,"a") as file_object:
			file_object.write(str(df_liveuamap.to_csv(file_txt,sep=" "))) 


		start_date=convert_string_to_date(string_start_date)
		start_date= start_date + timedelta(days=1)

		string_start_date=convert_date_to_string(start_date)

		i=i+1

	os.system('"'+file_txt+ '"')
	

def multiple_try(parameter1,parameter2,parameter3,nt,function):
	tries = nt
	for i in range(tries):
		try:
			function(parameter1,parameter2,parameter3)
		except (SSLError,ConnectionError) :
			if i < tries - 1: # i is zero indexed
				time.sleep(2)
				continue
			else:
				raise
		break


###########################################################################################

####
current_directory = os.getcwd()
extension="xlsx"

folder="Liveuamap"
directory_excel = os.path.join(current_directory, folder)

if not os.path.exists(directory_excel):
	os.makedirs(directory_excel)

file_excel=directory_excel+'\\Liveuamap_Scrapper.xlsx'
file_excel=file_excel.replace("\\", "\\\\")

file_txt=directory_excel+'\\Liveuamap_Scrapper.txt'
file_txt=file_txt.replace("\\", "\\\\")

###

ntry=50

today=datetime.now()
today=str(today)
today = re.sub(INVALID_TITLE_REGEX, '-', today)
print(today)


date_start_war = datetime(2022, 2, 24).date()
string_date_start_war = datetime.strftime(date_start_war, '%d.%m.%Y')
print(date_start_war) 


aujour=datetime.now().date()
string_aujour= aujour.strftime('%d.%m.%Y') 
print(string_aujour)

diff=aujour-date_start_war
diff=diff.days
print(diff)




url_liveuamap="https://liveuamap.com"
url_liveuamap_date="https://liveuamap.com/en/time/"


if sys.platform == "win32":
	multiple_try(string_date_start_war,diff,url_liveuamap_date,ntry,excel_liveuamap)
	#excel_liveuamap(string_date_start_war,diff,url_liveuamap_date)
else: 
	multiple_try(string_date_start_war,diff,url_liveuamap_date,ntry,file_liveuamap)
	#file_liveuamap(string_date_start_war,diff,url_liveuamap_date)


