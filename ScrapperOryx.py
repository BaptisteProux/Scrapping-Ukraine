import requests as r
from bs4 import BeautifulSoup as bs
import pandas as pd
import re
from itertools import chain
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os
import sys
from openpyxl.workbook.child import INVALID_TITLE_REGEX
import datetime
from os.path import exists
import openpyxl
import xlsxwriter as xs





#############################################################
##Fonctions
#Fonctions de nettoyage

def check_not_double_point (string):
		if string==":": 
			return False
		return True

def check_purge (string):
		if "Russia" in string or "Ukraine" in string or "-" in string or "of" in string or "which:" in string or "," in string: 
			return False
		return True

def convert_to_int(liste,integer=1):

	if integer==1:
		liste=[eval(i) if isinstance(i,str)==True else i for i in liste]

	else:
		j=0
		while j<integer:
			liste[j]=[eval(i) if isinstance(i,str)==True else i for i in liste[j]]
			j=j+1

	return liste

def applanir(liste,integer=1):
	i=0
	while i<integer:
		liste=list(chain.from_iterable(liste))
		i=i+1
	return liste

def harmonisation(equip_russe,equip_ukraine,list_russe,list_ukraine,data_russe,data_ukraine,length_type):

	equip_return=[]
	final_harmonisation=[]
	missing_ukraine=list(set(equip_russe)-set(equip_ukraine))
	missing_russe=list(set(equip_ukraine)-set(equip_russe))

	if len(missing_ukraine)!=0:
		for i in range(len(list_russe)):
			for j in range(len(missing_ukraine)):
				if equip_russe[i]==missing_ukraine[j]:
					equip_ukraine.insert(i,equip_russe[i])
					list_ukraine.insert(i,0)
					data_ukraine[0].insert(i,["0"])
					data_ukraine[1].insert(i,["0"]*length_type)				

	if len(missing_russe)!=0:
		for i in range(len(list_ukraine)):
			for j in range(len(missing_russe)):
				if equip_ukraine[i]==missing_russe[j]:
					equip_russe.insert(i,equip_ukraine[i])
					list_russe.insert(i,0)
					data_russe[0].insert(i,["0"])
					data_russe[1].insert(i,["0"]*length_type)	

	equip_return=equip_ukraine #peut aussi être equip_russe

	final_harmonisation.append(equip_return)
	final_harmonisation.append(list_russe)
	final_harmonisation.append(list_ukraine)
	final_harmonisation.append(data_russe)
	final_harmonisation.append(data_ukraine)


	return final_harmonisation

def nettoyage_encodage(categorie):
	vehicule_all=[]
	for i in range(len(categorie)):
		vehicule=categorie[i].get_text().split("\xa0")
		vehicule_all.append(vehicule)
	return vehicule_all

#Fonctions de récuperation de données

def parser(url,string="",classe="",parser="html.parser"):
	page=r.get(url)
	parser=bs(page.content,parser)
	parser.encoding = parser.apparent_encoding
	if string=="" :
		return parser
	else:
		if classe=="":
			sous_section=parser.find_all(str(string))
		else:
			sous_section=parser.find_all(str(string),{"class": classe})
		return sous_section

#Fonctions de traitement de données

def get_type_equipement(categorie):

	etat_type=[]
	donnée=[]

	for k in range(1,len(categorie)):
		if not "\n" in categorie[k]:

				etat=categorie[k].get_text().partition("(")[0]
				etat_all=etat.split(":")
				if any("Radars" in s for s in etat_all) and etat_all!=["Radars "]:  #Harmonisation nom categorie Radar entre les 2 listes
					etat_all=["Radars "] 
				if any("Unmanned Aerial Vehicles" in s for s in etat_all) and etat_all!=["Unmanned Aerial Vehicles "]:
					etat_all=["Unmanned Aerial Vehicles "] #Harmonisation nom categorie drone simple entre les 2 listes
				#if any("\xa0" in s for s in etat_all):
					#print(etat_all[0])
					#etat_all[0].replace("\xa0","")
				etat_type.append(etat_all)
	donnée.append(etat_type)

	donnée=applanir(donnée,2)

	return donnée


def scrap_donnee_equipement(categorie):

	sous_data=[]
	donnée=[]

	for k in range(1,len(categorie)):
		if not "\n" in categorie[k]:
				data=categorie[k].get_text().partition("(")[2]
				sous_data.append(data)
	donnée=sous_data
	return donnée



def get_data(categorie):

	donnée=[]
	data=[]
	totaux=[]
	sous_data=scrap_donnee_equipement(categorie)

	for i in range(len(sous_data)):
		current_data=sous_data[i].partition(":")
		current_data=list(filter(check_not_double_point,current_data))

		for j in range(len(current_data)):
			sous_current_data=current_data[j]
			sous_current_data=re.findall(r'\d+',sous_current_data)
			data.append(sous_current_data)

			

	for i in range(len(data)):
		if i%2==0:
			totaux.append(data[i])
			data[i]=":"

	data=list(filter(check_not_double_point,data))


	for i in range(len(data)):
		current_data=data[i]
		for j in range(len(current_data)):
			l=len(current_data)

			if l!=4:
				while l<4:
					current_data.insert(l-1,0)  #Harmonisation des données car Oryx n'a pas d'entrée si il y a 0 pertes repertorié dans une catégorie: pas forcement exacte en fonction des catégories mais privilige les catégories détruit et capturé car beaucoup plus de chance que c'est 2 cas soit concerné par des valeurs autre que 0 que les catégories endommagé et abandonné
					l=l+1
	donnée.append(totaux)
	donnée.append(data)
	return donnée



def get_total(categorie):

	total_value=[]
	total_type=[["Losses:"]]

	list_total=[]

	recap_totaux=categorie[0]
	recap_totaux=recap_totaux.get_text().partition("(")[0]

	recap_totaux=recap_totaux.split()

	for i in range(len(recap_totaux)):
		total=recap_totaux[i]
		total=re.findall(r'\d+',total)
		total_value.append(total)

	total_value=list(filter(None,total_value))

	for i in range(len(recap_totaux)):
		total=recap_totaux[i]
		total=re.findall(r'\D+',total)
		total_type.append(total)

	total_type=list(filter(check_purge,total_type))
	total_type=list(filter(None,total_type))

	list_total.append(total_type)
	list_total.append(total_value)

	return list_total

def get_vehicule(categorie):
	vehicule_all=[]
	for i in range(len(categorie)):
		engin=categorie[i]
		for j in range(len(engin)):
			vehicule=engin[j]
			if not any("(" in s for s in vehicule):
				vehicule_all.append(vehicule)
	return vehicule_all


def dataframe_converter(data,index,columns,astype="",obj=1):

	if obj==1:
		dataframe=pd.DataFrame(data,index,columns,dtype=object)
	else:
		dataframe=pd.DataFrame(data,index,columns)
	if astype!="":
		dataframe=dataframe.astype(str(astype))

	return dataframe

def ratio_between(data_Russe,data_Ukraine,luk=1):

	if luk==1:
		if len(data_Russe)!=len(data_Ukraine):
		 	print("Pas la même taille")
		 	return len(data_Russe)-len(data_Ukraine)
		else:
			ratio=[dr/du if du!=0 else np.NaN for dr,du in zip(data_Russe,data_Ukraine)]
			return ratio
	else: 
		j=0
		ratio=[[] for i in range(luk)]
		while j<luk:
			ratio[j]=[dr/data_Russe[j] if data_Russe[j]!=0 else np.NaN for dr in data_Ukraine[j]]
			j=j+1
		return ratio

def proportion(data):
	proportion=[]
	total=data[0]
	for i in range(1,len(data)):
		proportion.append(data[i]/total)
	return proportion


def get_image(url):
	link_list=[]
	page_russe=parser(url)
	page_russe=page_russe.find_all('a', href=True)
	for link in page_russe:
		if "damaged" in str(link) or "captured" in str(link)  or "destroyed" in str(link)  or "abandoned" in str(link)  : 
			if "jpg" in str(link) or "png" in str(link):
				link_list.append(str(link['href']))  #link['href']
	return link_list

def remove_split(liste,texte):
	for i in range(len(liste)):
		liste[i]=liste[i].split("texte")
	liste=applanir(liste)
	return liste



def texte(liste):
	texte=[]
	equip=[]
	for i in range(len(liste)):
		lis=liste[i].get_text().encode("utf-8")
		lis=str(lis).split("xa0")
		texte.append(lis)

	texte=applanir(texte)

	for i in range(len(texte)):
		if any("(" in s for s in texte[i]):
			text=texte[i].split("\\xc2\\")
			equip.append(text)
    
	equip=applanir(equip)


	for i in range(len(equip)):
		if any("(" in s for s in equip[i]):
			equip[i]=equip[i].split(":")
    
	equip=applanir(equip)

	for i in range(len(equip)):
			equip[i]=equip[i].split("b'")
    
	equip=applanir(equip)

	for i in range(len(equip)):
			equip[i]=equip[i].split('b"')
    
	equip=applanir(equip)



	for i in range(len(equip)):
		if not any("-" in s for s in equip[i]) or any("("in s for s in equip[i]) or any("x"in s for s in equip[i]) or any("Kartograf"in s for s in equip[i]):
			equip[i]=""

	equip=list(filter(None,equip))


	return equip

def csv_creator(dataframe,index=""):
	name =[x for x in globals() if globals()[x] is dataframe][0]

	file_csv=directory_oryx+'\\'+name+'.csv'

	file_csv=file_csv.replace("\\", "\\\\")

	if index!=False:
		dataframe.to_csv(file_csv)
	else:
		dataframe.to_csv(file_csv,index=False)






#############################################################
##

today=datetime.datetime.now()
today=str(today)
today = re.sub(INVALID_TITLE_REGEX, '-', today)
print(today)

#Parser
url_russie="https://www.oryxspioenkop.com/2022/02/attack-on-europe-documenting-equipment.html"
sous_section_russe=parser(url_russie,"h3")

image_russie=get_image(url_russie)

type_precis_equipement_russe=parser(url_russie,"li")
type_precis_equipement_russe=texte(type_precis_equipement_russe)



detail_equipement_russe=parser(url_russie,"ul")
equipement_type_russe=nettoyage_encodage(detail_equipement_russe[2:27])
equipement_type_russe=get_vehicule(equipement_type_russe)
vehicule_russe=pd.DataFrame(equipement_type_russe)


url_ukraine="https://www.oryxspioenkop.com/2022/02/attack-on-europe-documenting-ukrainian.html"
sous_section_ukraine=parser(url_ukraine,"h3")

image_ukraine=get_image(url_ukraine)

type_precis_equipement_ukraine=parser(url_ukraine,"li")
type_precis_equipement_ukraine=texte(type_precis_equipement_ukraine)

detail_equipement_ukraine=parser(url_ukraine,"ul")
equipement_type_ukraine=nettoyage_encodage(detail_equipement_ukraine[2:26])
equipement_type_ukraine=get_vehicule(equipement_type_ukraine)
vehicule_ukraine=pd.DataFrame(equipement_type_ukraine)

#Recuperation data
data_russe=get_data(sous_section_russe)
totaux_russe=data_russe[0]
totaux_russe=applanir(totaux_russe)
totaux_russe=convert_to_int(totaux_russe)
type_equipement_russe=get_type_equipement(sous_section_russe)
total_russe=get_total(sous_section_russe)

type_total_russe=applanir(total_russe[0])
value_total_russe=applanir(total_russe[1])
value_total_russe=convert_to_int(value_total_russe)

data_ukraine=get_data(sous_section_ukraine)
totaux_ukraine=data_ukraine[0]
totaux_ukraine=applanir(totaux_ukraine)
totaux_ukraine=convert_to_int(totaux_ukraine)
type_equipement_ukraine=get_type_equipement(sous_section_ukraine)
total_ukraine=get_total(sous_section_ukraine)

type_total_ukraine=applanir(total_ukraine[0])
value_total_ukraine=applanir(total_ukraine[1])
value_total_ukraine=convert_to_int(value_total_ukraine)


####

current_directory = os.getcwd()

folder="oryx"
directory_oryx = os.path.join(current_directory, folder)

if not os.path.exists(directory_oryx):
	os.makedirs(directory_oryx)

###




#Harmonisation
if(set(type_total_ukraine)==set(type_total_russe)):
	type_total=type_total_russe

length_type=len(type_total)-1

if (set(type_equipement_ukraine)!=set(type_equipement_russe)):
	data_harmonie=harmonisation(type_equipement_russe,type_equipement_ukraine,totaux_russe,totaux_ukraine,data_russe,data_ukraine,length_type)
	type_equipement=data_harmonie[0]
	totaux_russe=data_harmonie[1]
	totaux_ukraine=data_harmonie[2]
	data_russe=data_harmonie[3]
	data_ukraine=data_harmonie[4]
else:
	type_equipement=type_equipement_russe

ratio_total_russe_ukraine=ratio_between(value_total_russe,value_total_ukraine)
ratio_totaux_russe_ukraine=ratio_between(totaux_russe,totaux_ukraine)

data_russe_int=convert_to_int(data_russe[1],len(data_russe[1]))
data_ukraine_int=convert_to_int(data_ukraine[1],len(data_ukraine[1]))

pourcentage_equipement_russe=ratio_between(totaux_russe,data_russe_int,len(data_russe_int))
pourcentage_equipement_ukraine=ratio_between(totaux_ukraine,data_ukraine_int,len(data_ukraine_int))


#Convertit en dataframe pour l'excel
value_total=[]
value_total.append(value_total_russe)
value_total.append(value_total_ukraine)


ratio_total=[]+value_total
ratio_total.append(ratio_total_russe_ukraine)


data_total_excel=dataframe_converter(ratio_total,["Russie","Ukraine","Ratio"],type_total)


print(data_total_excel)

proportion_russe=proportion(value_total_russe)
proportion_russe_excel=pd.DataFrame(proportion_russe,type_total[1:5],["Proportion vehicules par categorie de perte"])
print(proportion_russe_excel)

proportion_ukraine=proportion(value_total_ukraine)
proportion_ukraine_excel=pd.DataFrame(proportion_ukraine,type_total[1:5],["Proportion vehicules par categorie de perte"])
print(proportion_ukraine_excel)


totaux=[]
totaux.append(totaux_russe)
totaux.append(totaux_ukraine)


ratio_totaux=[]+totaux
ratio_totaux.append(ratio_totaux_russe_ukraine)



data_totaux_excel=dataframe_converter(ratio_totaux,["Russie","Ukraine","Ratio"],type_equipement)
print(data_totaux_excel)



data_per_type_russe=dataframe_converter(data_russe[1],type_equipement,type_total[1:],"int")
print(data_per_type_russe)


pourcentage_russe=[]
pourcentage_russe.append(pourcentage_equipement_russe)

data_pourcentage_russe=pd.DataFrame(pourcentage_russe,["Pourcentage"],type_equipement)
print(data_pourcentage_russe)


data_per_type_ukraine=dataframe_converter(data_ukraine[1],type_equipement,type_total[1:],"int")
print(data_per_type_ukraine)

pourcentage_ukraine=[]
pourcentage_ukraine.append(pourcentage_equipement_ukraine)

data_pourcentage_ukraine=pd.DataFrame(pourcentage_ukraine,["Pourcentage"],type_equipement)
print(data_pourcentage_ukraine)


data_per_type=pd.concat([data_per_type_russe, data_per_type_ukraine], axis=1)
print(data_per_type)


data_type_precis_equipement_russe=pd.DataFrame(type_precis_equipement_russe)
data_type_precis_equipement_ukraine=pd.DataFrame(type_precis_equipement_ukraine)

print(data_type_precis_equipement_russe)
print(data_type_precis_equipement_ukraine)



data_total_excel=pd.concat([data_total_excel],keys=['Total '+today])
data_type_precis_equipement_russe=pd.concat([data_type_precis_equipement_russe],keys=['type russe '+today])
data_type_precis_equipement_ukraine=pd.concat([data_type_precis_equipement_ukraine],keys=['type ukraine '+today])
data_totaux_excel=pd.concat([data_totaux_excel],keys=['Total '+today])
data_proportion_all=pd.concat([proportion_russe_excel,proportion_ukraine_excel],keys=['Prop russes '+today,'Prop ukrainiens '+today])
data_per_type_all=pd.concat([data_per_type_russe,data_per_type_ukraine],keys=['detail russe '+today,'detail ukraine '+today])
data_pourcentage_all=pd.concat([data_pourcentage_russe,data_pourcentage_ukraine],keys=['% russe '+today,'% ukraine '+today])


csv_creator(data_total_excel)
csv_creator(data_type_precis_equipement_russe)
csv_creator(data_type_precis_equipement_ukraine)
csv_creator(data_totaux_excel)
csv_creator(data_proportion_all)
csv_creator(data_per_type_all)
csv_creator(data_pourcentage_all)



file_text=directory_oryx+'\\Data_Oryx.txt'

file_text=file_text.replace("\\", "\\\\")

all_dataframe=pd.concat([data_total_excel,data_totaux_excel,data_proportion_all,data_per_type_all,data_pourcentage_all])
with open(file_text,"a") as file_object:
	file_object.write(str(all_dataframe.to_csv(file_text,sep=" ")))   #Ne fais pas ce que je veux : ,n'append pas mais overwrite 


file_excel=directory_oryx+'\\scrapperOryx_to_excel.xlsx'

file_excel=file_excel.replace("\\", "\\\\")

if sys.platform == "win32" and exists(file_excel)==True:

	with pd.ExcelWriter(file_excel,engine="openpyxl",mode="a", if_sheet_exists="overlay") as writer:
		data_total_excel.to_excel(writer, sheet_name='Total ',startrow=writer.sheets['Total '].max_row)
		data_proportion_all.to_excel(writer, sheet_name='Proportion ',startrow=writer.sheets['Proportion '].max_row)
		data_totaux_excel.to_excel(writer, sheet_name='totaux ',startrow=writer.sheets['totaux '].max_row)
		data_per_type_all.to_excel(writer, sheet_name='detail ',startrow=writer.sheets['detail '].max_row)
		data_pourcentage_all.to_excel(writer, sheet_name='% ',startrow=writer.sheets['% '].max_row)
		#data_type_precis_equipement_russe.to_excel(writer, sheet_name='type russe ',startrow=writer.sheets['% '].max_row)
		#data_type_precis_equipement_ukraine.to_excel(writer, sheet_name='type ukraine ',startrow=writer.sheets['% '].max_row)
		#data_per_type.to_excel(writer, sheet_name='detail all')
		#vehicule_russe.to_excel(writer,sheet_name="vehicules perdus totaux russes "+today)
		#vehicule_ukraine.to_excel(writer,sheet_name="vehicules perdus totaux ukraine "+today)
	os.system('"'+file_excel+ '"')


elif sys.platform == "win32" and exists(file_excel)!=True:

	with pd.ExcelWriter(file_excel) as writer:
		data_total_excel.to_excel(writer, sheet_name='Total ')
		data_proportion_all.to_excel(writer, sheet_name='Proportion ')
		data_totaux_excel.to_excel(writer, sheet_name='totaux ')
		data_per_type_all.to_excel(writer, sheet_name='detail ')
		data_pourcentage_all.to_excel(writer, sheet_name='% ')
		data_type_precis_equipement_russe.to_excel(writer, sheet_name='type russe ')
		data_type_precis_equipement_ukraine.to_excel(writer, sheet_name='type ukraine ')
	os.system('"'+file_excel+ '"')

else:
	os.system(file_text)

#plot
plt.figure(1)


X_axix=np.arange(len(type_equipement))

plt.xticks(fontsize=5, rotation=90)
plt.xticks(X_axix,type_equipement)


plt.plot(X_axix,totaux_russe,c="red",label="losses russian")
plt.plot(X_axix,totaux_ukraine,c="blue",label="losses ukrainian")
plt.title("Graphique pertes totaux")
plt.xlabel("type of vehicle")
plt.ylabel("number")



for i,j in zip(X_axix,totaux_russe):
	plt.annotate(str(j),xy=(i,j),xytext=(5,5),textcoords="offset points",fontweight="bold",color="red",fontsize=5)

for i,j in zip(X_axix,totaux_ukraine):
	plt.annotate(str(j),xy=(i,j),xytext=(5,5),textcoords="offset points",fontweight="bold",color="blue",fontsize=5)


file_image=directory_oryx+'\\Graph_loses_totaux.png'

file_image=file_image.replace("\\", "\\\\")

plt.savefig(file_image)

os.startfile(file_image)

plt.figure(2)

plt.pie(proportion_russe,labels=type_total[1:],autopct="%.2f%%")
plt.title("Proportion types de pertes vehicules russes")

file_image=directory_oryx+'\\Pie_types_de_pertes_russe.png'

file_image=file_image.replace("\\", "\\\\")

plt.savefig(file_image)
os.startfile(file_image)

plt.figure(3)

plt.pie(proportion_ukraine,labels=type_total[1:],autopct="%.2f%%")
plt.title("Proportion types de pertes vehicules ukrainiens")

file_image=directory_oryx+'\\Pie_types_de_pertes_ukraine.png'

file_image=file_image.replace("\\", "\\\\")

plt.savefig(file_image)

os.startfile(file_image)




plt.figure(4)

plt.bar(X_axix -0.2,height=totaux_russe,width=0.5,color="r",label = "russe")
plt.bar(X_axix + 0.2,height=totaux_ukraine,width=0.5,color="b",label = "ukrainien")
plt.title("Graphique en bar perte totaux")


plt.xticks(fontsize=5, rotation=90)
plt.xticks(X_axix,type_equipement)

for i,j in enumerate(totaux_russe):
	plt.text(i , j , str(j),fontweight="bold",color="red",fontsize=5)

for i,j in enumerate(totaux_ukraine):
	plt.text(i , j , str(j),fontweight="bold",color="blue",fontsize=5)


file_image=directory_oryx+'\\Graph_loses_bar_totaux.png'

file_image=file_image.replace("\\", "\\\\")

plt.savefig(file_image)

os.startfile(file_image)



plt.figure(5)

X_axix_total=np.arange(len(type_total))


plt.xticks(X_axix_total,type_total)

plt.bar(X_axix_total -0.2,height=value_total_russe,width=0.5,color="r",label = "russe")
plt.bar(X_axix_total + 0.2,height=value_total_ukraine,width=0.5,color="b",label = "ukraine")

for i,j in enumerate(value_total_russe):
	plt.text(i , j , str(j),fontweight="bold",color="red",fontsize=5)

for i,j in enumerate(value_total_ukraine):
	plt.text(i , j , str(j),fontweight="bold",color="blue",fontsize=5)

plt.title("Graphique en bar pertes total")

file_image=directory_oryx+'\\Graph_loses_bar_total.png'

file_image=file_image.replace("\\", "\\\\")

plt.savefig(file_image)

os.startfile(file_image)
'''
os.startfile("Pie_types_de_pertes_russe.png")
os.startfile("Pie_types_de_pertes_ukraine.png")
os.startfile("Graph_loses_totaux.png")
os.startfile("Graph_loses_bar_totaux.png")
#plt.show()
'''